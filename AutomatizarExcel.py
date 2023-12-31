# Sistema automatizado de reportes (Jesus Morillo)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string

archivo_excel = pd.read_excel('supermarket_sales.xlsx')
print(archivo_excel[['Gender', 'Product line', 'Total']])

tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)

tabla_pivote.to_excel('sales_2023.xlsx', startrow=4, sheet_name='Report')

# Aquí comienza openpyxl

wb = load_workbook('sales_2023.xlsx')
pestaña = wb['Report']

# Aquí detecta el mínimo de columnas activas y máximas

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

# Gráfico 

barchart = BarChart()

data = Reference(pestaña, min_col=min_col+1, max_col=max_col, min_row=min_fila, max_row=max_fila)
categorias = Reference(pestaña, min_col=min_col, max_col=min_col, min_row=min_fila+1, max_row=max_fila)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categorias)

pestaña.add_chart(barchart, 'B12')
barchart.title = 'Ventas'  # Aquí se coloca = para asignar el valor no " (Ventas) "
barchart.style = 2
# Aquí va fórmula de suma

abecedario = list(string.ascii_uppercase)
abecedario_excel = (abecedario[0:max_col])

for i  in abecedario_excel:
    if i!='A':
        pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
        pestaña[f'{i}{max_fila+1}'].style = 'Currency'

pestaña[f'{abecedario_excel[0]}{max_fila+1}'] = 'Total'

pestaña['A1'] = 'Reporte ventas (Jesus Morillo)'
pestaña['A2'] = '2023'

pestaña['A1'].font = Font('Arial', bold=True, size=20)
pestaña['A2'].font = Font('Arial', bold=True, size=14)

wb.save('sales_2023.xlsx')

